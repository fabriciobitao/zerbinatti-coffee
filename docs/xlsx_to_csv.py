"""Converte xlsx multi-sheet em CSV consolidado pra subir no Drive como Google Sheet."""
from openpyxl import load_workbook
import csv
import io

src = r"C:\Users\fabio\dev\zerbinatti-coffee\docs\custos-plataforma-2026-05-08.xlsx"
out = r"C:\Users\fabio\dev\zerbinatti-coffee\docs\custos-plataforma-2026-05-08.csv"

wb = load_workbook(src, data_only=False)
buf = io.StringIO()
w = csv.writer(buf)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    w.writerow([f"=== ABA: {sheet} ==="])
    for row in ws.iter_rows(values_only=True):
        # Resolve formulas as values where possible
        clean = []
        for c in row:
            if c is None:
                clean.append("")
            else:
                clean.append(str(c))
        w.writerow(clean)
    w.writerow([])  # linha em branco entre abas
    w.writerow([])

content = buf.getvalue()

with open(out, "w", encoding="utf-8-sig", newline="") as f:
    f.write(content)

print(f"Saved: {out} ({len(content)} chars)")
