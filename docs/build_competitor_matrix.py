"""
Gera planilha XLSX comparativa: Zerbinatti vs 12 concorrentes café especial BR.
Origem dos dados: docs/CONCORRENTES-CAFE-ESPECIAL-2026-05-07.md + análise do código atual do site.
Convenção: ✓ = tem o feature, ✗ = não tem, "—" = N/D, valores literais quando aplicável.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = "/Users/fabricio/dev/cafe/docs/CONCORRENTES-CAFE-ESPECIAL-2026-05-07.xlsx"

# Ordem de colunas: Zerbinatti primeiro, depois concorrentes em ordem do relatório
# (nome, url do site)
BRANDS_WITH_URL = [
    ("Zerbinatti",       "https://zerbinatti-coffee-259156177034.southamerica-east1.run.app/"),
    ("Coffee Lab",       "https://loja.coffeelab.com.br/"),
    ("Octavio Café",     "https://octaviocafe.com.br/"),
    ("Suplicy",          "https://www.suplicycafes.com.br/"),
    ("Orfeu",            "https://www.cafeorfeu.com.br/"),
    ("Daterra",          "https://daterracoffee.com.br/"),
    ("Wolff",            "https://www.wolffcafe.com.br/"),
    ("Lucca",            "https://www.luccacafesespeciais.com.br/"),
    ("UM Coffee Co",     "https://www.umcoffeeco.com/"),
    ("3 Corações",       "https://www.3coracoes.com.br/marca/cafes-especiais/"),
    ("Academia do Café", "https://www.academiadocafe.com.br/"),
    ("Moka Clube",       "https://www.mokaclube.com.br/"),
    ("Coffee++",         "https://coffeemais.com/"),
]
BRANDS = [b[0] for b in BRANDS_WITH_URL]
BRAND_URLS = {b[0]: b[1] for b in BRANDS_WITH_URL}

# Cada feature: (categoria, feature, [valor_por_marca_na_ordem_acima])
# Use "✓" / "✗" / "—" (N/D) ou texto literal (preços, prazos, etc).

# Para Zerbinatti, valores derivados do código atual (page.tsx, Subscription.tsx,
# Products.tsx, BestSellers.tsx, TrustBar.tsx, B2B.tsx, entregas/page.tsx, cafes/[slug]/page.tsx).

ROWS = [
    # ============= 1. CATÁLOGO DE PRODUTO =============
    ("CATÁLOGO DE PRODUTO", "SKUs visíveis (qtd)",
     ["3", "~15", "~10", "—", "67", "~50+", "4", "20", "~15", "4", "~15", "~30", "~10"]),
    ("CATÁLOGO DE PRODUTO", "Single Origin",
     ["✓", "✓", "✓", "—", "✓", "✓", "—", "✓", "✓", "—", "✓", "✓", "✓"]),
    ("CATÁLOGO DE PRODUTO", "Blends",
     ["✓", "✓", "✓", "—", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "—"]),
    ("CATÁLOGO DE PRODUTO", "Microlotes",
     ["✓", "✗", "✓", "—", "✓", "✓", "✗", "✓", "✓", "—", "✗", "✓", "✓"]),
    ("CATÁLOGO DE PRODUTO", "Nanolotes",
     ["✗", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✓", "✗", "✗", "✗", "✗"]),
    ("CATÁLOGO DE PRODUTO", "Cápsulas",
     ["✗", "✓", "✓", "—", "✓", "✗", "✓", "✗", "✗", "✓", "✗", "✗", "✓"]),
    ("CATÁLOGO DE PRODUTO", "Drip coffee / ESE pods",
     ["✗", "✗", "✗", "—", "✓", "✗", "✓", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("CATÁLOGO DE PRODUTO", "Café em grão verde (B2B)",
     ["✓", "✗", "✗", "—", "✗", "✓", "✗", "✓", "✗", "✗", "✗", "✗", "✗"]),
    ("CATÁLOGO DE PRODUTO", "Acessórios (moedor, bule, xícara)",
     ["✗", "✗", "✗", "—", "✓", "✗", "✗", "✗", "✗", "✗", "✓", "✗", "✗"]),
    ("CATÁLOGO DE PRODUTO", "Kits / Presentes",
     ["✗", "✗", "✗", "—", "✓", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),

    # ============= 2. EMBALAGENS =============
    ("EMBALAGENS", "250g",
     ["✗", "✓", "✓", "—", "✓", "✓", "✓", "✓", "✓", "—", "✓", "✓", "✓"]),
    ("EMBALAGENS", "500g",
     ["✓", "✗", "✓", "—", "✗", "✗", "✓", "✗", "✗", "—", "✗", "✗", "✗"]),
    ("EMBALAGENS", "1kg",
     ["✗", "✗", "✗", "—", "✓", "✗", "✓", "✗", "✗", "—", "✗", "✓", "✗"]),
    ("EMBALAGENS", "Sample / 50g",
     ["✗", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "—", "✗", "✓", "✗"]),
    ("EMBALAGENS", "Embalagem proprietária (ex: Penta Box)",
     ["✗", "✗", "✗", "—", "✓", "✓", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("EMBALAGENS", "Válvula degasificação",
     ["✓", "✓", "✓", "—", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"]),

    # ============= 3. MOAGEM =============
    ("MOAGEM", "Café em grão",
     ["✓", "✓", "✓", "—", "✓", "—", "✓", "✓", "✓", "—", "—", "✓", "✓"]),
    ("MOAGEM", "Moído (genérico)",
     ["✓", "✓", "✓", "—", "✓", "—", "✓", "✓", "✓", "—", "—", "✓", "✓"]),
    ("MOAGEM", "Moagem por método (V60, Moka, Prensa, Espresso)",
     ["✗", "✗", "✗", "—", "✗", "✗", "✓", "✓", "✗", "✗", "✗", "✓", "✗"]),

    # ============= 4. FICHA TÉCNICA =============
    ("FICHA TÉCNICA", "Score SCA na ficha do produto",
     ["✓ (84,75)", "✗", "✗", "—", "✗", "✓ (80–89+)", "—", "✗ (só ‘80+’)", "—", "✓ (85+)", "✗", "✓ (84+)", "—"]),
    ("FICHA TÉCNICA", "Altitude por SKU",
     ["✓", "✗", "✗", "—", "✗", "✓", "—", "✗", "—", "—", "✗", "✗", "—"]),
    ("FICHA TÉCNICA", "Variedade botânica",
     ["✓", "✗", "✗", "—", "✗", "✓", "—", "✗", "—", "—", "✗", "✗", "—"]),
    ("FICHA TÉCNICA", "Processo (natural, lavado, etc)",
     ["✓", "✗", "✗", "—", "✓", "✓", "—", "✓", "✓", "—", "✗", "✓", "—"]),
    ("FICHA TÉCNICA", "Safra/colheita",
     ["✓", "✗", "✗", "—", "✗", "✓", "—", "✗", "—", "—", "✗", "✗", "—"]),
    ("FICHA TÉCNICA", "Notas sensoriais (descritores)",
     ["✓", "✓", "✓", "—", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"]),
    ("FICHA TÉCNICA", "Gráfico radar de perfil sensorial",
     ["✓", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("FICHA TÉCNICA", "Métodos recomendados",
     ["✓", "✗", "✗", "—", "✗", "✓", "✗", "✓", "✗", "✗", "✗", "✓", "✗"]),
    ("FICHA TÉCNICA", "Nome do produtor / fazenda",
     ["✓", "✗", "✗", "—", "✗", "✓", "✗", "✓", "✓", "✗", "✗", "✓", "✗"]),
    ("FICHA TÉCNICA", "Hierarquia de linhas (premium tiers)",
     ["✓", "✗", "✗", "—", "✓", "✓", "✓", "✓", "✓", "✓", "✗", "✗", "✗"]),

    # ============= 5. PREÇO =============
    ("PREÇO", "Faixa de preço (250g equivalente)",
     ["R$ 35–60 (250g eq.)", "R$ 39–134", "—", "—", "R$ 42+", "—", "R$ 30–40", "R$ 59–129",
      "R$ 82–247", "—", "R$ 62–155", "R$ 68–129", "R$ 39–69 (est.)"]),

    # ============= 6. PAGAMENTOS =============
    ("PAGAMENTOS", "Pix exibido no site",
     ["✓ (-10% destaque)", "✗", "✗", "—", "✗", "—", "—", "—", "✗", "—", "—", "✗", "—"]),
    ("PAGAMENTOS", "Desconto Pix destacado",
     ["✓", "✗", "✗", "—", "✗", "—", "—", "—", "✗", "—", "—", "✗", "—"]),
    ("PAGAMENTOS", "Pix parcelado (Pagaleve etc)",
     ["✗", "✗", "✗", "—", "✗", "—", "—", "—", "✗", "—", "—", "✗", "—"]),
    ("PAGAMENTOS", "Cartão de crédito",
     ["✓", "✓", "✓", "—", "✓", "—", "—", "✓", "✓", "—", "—", "✓", "✓"]),
    ("PAGAMENTOS", "Parcelamento sem juros",
     ["✓ (4x)", "—", "—", "—", "✓ (3x)", "—", "—", "—", "—", "—", "—", "✓ (5x)", "—"]),
    ("PAGAMENTOS", "Boleto",
     ["✗", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"]),
    ("PAGAMENTOS", "Apple Pay / Google Pay",
     ["✗", "✗", "✗", "—", "✗", "—", "—", "✗", "✓", "—", "—", "✗", "—"]),
    ("PAGAMENTOS", "Carteira do cliente / Loja na conta",
     ["✗", "✗", "✗", "—", "✓ (VTEX)", "—", "—", "—", "✓", "✓ (Mercafé)", "—", "✓", "—"]),
    ("PAGAMENTOS", "Cupom primeira compra exibido",
     ["✗", "✗", "✗", "—", "✗", "—", "—", "—", "✗", "—", "—", "✓ (QUEROCAFE)", "—"]),
    ("PAGAMENTOS", "Programa fidelidade / cashback",
     ["✗", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("PAGAMENTOS", "Programa de pontos",
     ["✗", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),

    # ============= 7. FRETE / LOGÍSTICA =============
    ("FRETE / LOGÍSTICA", "Frete grátis (limiar)",
     ["R$ 99 (Brasil todo)", "—", "—", "—", "Por plano (R$ 175+)", "✗", "Via Club", "Via planos",
      "—", "—", "Varia por região", "R$ 199 (S/SE)", "Via Amazon Prime"]),
    ("FRETE / LOGÍSTICA", "Frete grátis Brasil inteiro",
     ["✓ (R$ 99+)", "—", "—", "—", "✓ (R$ 435+)", "✗", "—", "—", "—", "—", "✗", "✗", "—"]),
    ("FRETE / LOGÍSTICA", "Calendário fixo de envio",
     ["✓ (sex pós-quarta)", "✗", "✗", "—", "✓ (até dia 24)", "✗", "✗", "✗", "✗", "✗", "✗", "✓ (mensal)", "✗"]),
    ("FRETE / LOGÍSTICA", "Torra sob demanda comunicada",
     ["✓", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("FRETE / LOGÍSTICA", "Prazo declarado (capitais SP/RJ)",
     ["2–4 dias úteis", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"]),
    ("FRETE / LOGÍSTICA", "Rastreamento por e-mail/WhatsApp",
     ["✓", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"]),
    ("FRETE / LOGÍSTICA", "Política de troca/devolução pública",
     ["✓ (pág. dedicada)", "✗", "✗", "—", "—", "—", "✗", "✓", "—", "—", "—", "—", "—"]),
    ("FRETE / LOGÍSTICA", "Garantia de satisfação 1ª compra",
     ["✓ (1º mês)", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("FRETE / LOGÍSTICA", "QR code de rastreabilidade no pacote",
     ["✗", "✗", "✗", "—", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),

    # ============= 8. ASSINATURA =============
    ("ASSINATURA", "Tem clube de assinatura",
     ["✓", "✗", "—", "—", "✓", "✗ (B2B only)", "✓", "✓", "✓", "✓ (Mercafé)", "—", "✓", "—"]),
    ("ASSINATURA", "Tiers escalonados (3+ planos)",
     ["✓ (3 tiers)", "✗", "—", "—", "✓ (3 + Confraria)", "✗", "—", "✓ (3)", "✗ (1 ou 2 pkts)", "—", "—", "✓ (2)", "—"]),
    ("ASSINATURA", "Plano premium tipo ‘Confraria’ / Herdeiro",
     ["✓ (Herdeiro)", "✗", "—", "—", "✓ (Confraria)", "✗", "✗", "✓ (Sublime)", "✗", "—", "—", "✗", "—"]),
    ("ASSINATURA", "Pause/cancelar sem multa comunicado",
     ["✓", "—", "—", "—", "—", "—", "—", "✓", "✓", "—", "—", "✓", "—"]),
    ("ASSINATURA", "Customização de café (trocar lote)",
     ["✓ (Sommelier+)", "✗", "—", "—", "—", "—", "—", "✓", "—", "—", "—", "✓", "—"]),
    ("ASSINATURA", "Brinde de boas-vindas",
     ["✗", "✗", "—", "—", "✓ (xícaras)", "✗", "—", "—", "—", "—", "—", "—", "—"]),
    ("ASSINATURA", "Acesso antecipado a microlotes",
     ["✓ (Sommelier)", "—", "—", "—", "✓ (Confraria)", "—", "—", "✓", "—", "—", "—", "—", "—"]),
    ("ASSINATURA", "Bilhete escrito do torrador",
     ["✓", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("ASSINATURA", "Eventos exclusivos / visita à fazenda",
     ["✓ (Herdeiro)", "✗", "✗", "—", "✓ (Confraria)", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("ASSINATURA", "FAQ dedicado da assinatura",
     ["✓", "✗", "✗", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"]),

    # ============= 9. B2B =============
    ("B2B", "Página B2B dedicada",
     ["✓", "✓", "—", "✓ (Visa app)", "✓", "✓", "—", "✓", "—", "✓", "—", "—", "—"]),
    ("B2B", "Cafeterias / restaurantes",
     ["✓", "✓", "—", "✓", "✓", "✓", "—", "✓", "—", "✓", "—", "—", "—"]),
    ("B2B", "Programa para escritórios",
     ["✓", "✗", "—", "✗", "✓", "✗", "—", "✗", "—", "✓", "—", "—", "—"]),
    ("B2B", "Self-service B2B (checkout sem comercial)",
     ["✗", "✗", "—", "✗", "✗", "✗", "—", "✗", "—", "✗", "—", "—", "—"]),
    ("B2B", "NF-e automática",
     ["—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"]),
    ("B2B", "Café verde (grão cru) para torrefações",
     ["✓", "✗", "—", "✗", "✗", "✓", "✗", "✓", "✗", "✗", "✗", "✗", "✗"]),
    ("B2B", "Blend personalizado",
     ["✓", "✗", "—", "✗", "✗", "—", "✗", "—", "✗", "—", "✗", "✗", "✗"]),

    # ============= 10. STORYTELLING / MARCA =============
    ("STORYTELLING / MARCA", "Narrativa de fundador/produtor forte",
     ["✓ (Giuseppe 1897)", "✓ (Isabela Raposeiras)", "—", "—", "—", "—", "—", "✓ (Georgia Franco)",
      "✓", "—", "—", "—", "—"]),
    ("STORYTELLING / MARCA", "Storytelling histórico (>50 anos)",
     ["✓ (1897)", "✗", "—", "—", "✗", "✗", "✗", "✗", "✗", "✓ (3 Corações)", "✗", "✗", "✗"]),
    ("STORYTELLING / MARCA", "Página dedicada à fazenda",
     ["✓ (Valim)", "✗", "—", "—", "✗", "✓", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("STORYTELLING / MARCA", "Página dedicada ao processo",
     ["✓", "✗", "—", "—", "✗", "✓", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("STORYTELLING / MARCA", "Premiações Cup of Excellence comunicadas",
     ["✗", "—", "—", "—", "✓ (28x)", "✓", "—", "✗", "✓", "—", "—", "✗", "—"]),
    ("STORYTELLING / MARCA", "Selo B Corp / sustentabilidade ESG",
     ["✗", "✗", "✗", "—", "✗", "✓", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("STORYTELLING / MARCA", "Identidade visual italiana / herança",
     ["✓", "✗", "✓", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),

    # ============= 11. CONTEÚDO EDUCATIVO =============
    ("CONTEÚDO EDUCATIVO", "Blog / revista",
     ["✓ (Revista Zerbinatti)", "✓", "✓ (Saberes do Café)", "—", "✗", "✓ (News from Farm)",
      "✗", "✓", "✗", "✗", "✗", "✓", "✗"]),
    ("CONTEÚDO EDUCATIVO", "Cursos pagos / escola",
     ["✗", "✓ (R$ 1.859–5.020)", "✗", "—", "✗", "✗", "✗", "✓ (Lucca Lab)", "✗", "✗", "✗", "✗", "✗"]),
    ("CONTEÚDO EDUCATIVO", "Vídeo embedded de preparo na ficha",
     ["✗", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("CONTEÚDO EDUCATIVO", "Guia de moagem por método",
     ["✗", "✓", "—", "—", "—", "✓", "—", "✓", "—", "—", "✗", "✓", "—"]),
    ("CONTEÚDO EDUCATIVO", "Quiz para descobrir café ideal",
     ["✓ (Quiz.tsx)", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),

    # ============= 12. PROVA SOCIAL / REVIEWS =============
    ("PROVA SOCIAL / REVIEWS", "Reviews de clientes na ficha",
     ["✓ (com texto + autor + método)", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("PROVA SOCIAL / REVIEWS", "Estrelas (rating médio)",
     ["✓", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("PROVA SOCIAL / REVIEWS", "Filtro de reviews por descritor sensorial",
     ["✗", "✗", "✗", "—", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("PROVA SOCIAL / REVIEWS", "Mídia espontânea (imprensa) destacada",
     ["✗", "✓", "✓", "—", "✓", "✓", "✗", "✓", "✗", "✓", "✗", "✗", "✗"]),

    # ============= 13. UX / TÉCNICO =============
    ("UX / TÉCNICO", "Plataforma",
     ["Custom Next.js/Cloud Run", "Iluria (legado)", "—", "—", "VTEX", "WordPress (B2B)", "—",
      "WooCommerce", "Shopify", "Mercafé", "Wix", "Shopify", "Shopify"]),
    ("UX / TÉCNICO", "Mobile-first responsivo",
     ["✓", "—", "—", "—", "—", "—", "—", "—", "✓", "—", "✗", "✓", "✓"]),
    ("UX / TÉCNICO", "Performance/SSR moderno",
     ["✓", "✗", "✗", "—", "—", "—", "—", "✗", "✓", "—", "✗", "✓", "✓"]),
    ("UX / TÉCNICO", "Cart drawer / carrinho lateral",
     ["✓", "—", "—", "—", "✓", "—", "—", "—", "✓", "—", "—", "✓", "✓"]),
    ("UX / TÉCNICO", "Páginas de produto com slug SEO",
     ["✓", "—", "—", "—", "✓", "—", "—", "✓", "✓", "—", "—", "✓", "—"]),
    ("UX / TÉCNICO", "Breadcrumb",
     ["✓", "—", "—", "—", "✓", "—", "—", "—", "✓", "—", "—", "✓", "—"]),
    ("UX / TÉCNICO", "Produtos relacionados na ficha",
     ["✓", "—", "—", "—", "✓", "—", "—", "—", "✓", "—", "—", "✓", "—"]),
    ("UX / TÉCNICO", "Scroll reveal / animações",
     ["✓", "✗", "✗", "—", "—", "—", "✗", "—", "—", "—", "✗", "—", "—"]),

    # ============= 14. INTERNACIONAL =============
    ("INTERNACIONAL", "Multi-idioma (PT/EN/ES)",
     ["✗", "✗", "✗", "—", "✗", "✓ (PT/EN/ES/JP)", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("INTERNACIONAL", "Checkout em moeda estrangeira",
     ["✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
    ("INTERNACIONAL", "Envio internacional (DTC)",
     ["✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),

    # ============= 15. CONTATO / CANAIS =============
    ("CONTATO / CANAIS", "WhatsApp direto na home/ficha",
     ["✓", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"]),
    ("CONTATO / CANAIS", "Newsletter (captura de e-mail)",
     ["✓", "✓", "✓", "—", "✓", "✓", "—", "✓", "✓", "✓", "—", "✓", "✓"]),
    ("CONTATO / CANAIS", "Instagram link",
     ["✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓", "✓"]),
    ("CONTATO / CANAIS", "YouTube / TikTok link",
     ["✓", "✓", "✗", "—", "✓", "✗", "✗", "—", "✗", "✓", "✗", "—", "—"]),
    ("CONTATO / CANAIS", "App mobile próprio",
     ["✗", "✗", "✗", "✓ (Visa)", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗"]),
]

# === Construção do workbook ===
wb = Workbook()
ws = wb.active
ws.title = "Comparativo Master"

# Cores
COLOR_HEADER = "1A1108"        # coffee-950
COLOR_HEADER_FONT = "F4E4C1"   # gold-100
COLOR_CATEGORY = "8B6F47"      # coffee-600
COLOR_NOSSO = "2F5233"         # verde escuro destacar Zerbinatti
COLOR_OK = "C8E6C9"            # verde claro
COLOR_NO = "FFCDD2"            # vermelho claro
COLOR_ND = "ECEFF1"            # cinza claro
COLOR_ALT = "FAF7F2"           # bege alternado

# Fontes/bordas
header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
nosso_font = Font(bold=True, color="FFFFFF", size=11)
category_font = Font(bold=True, color="FFFFFF", size=10)
thin = Side(border_style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# === Linha 1: Título ===
ws.cell(row=1, column=1, value="Café Especial — Matriz Comparativa Zerbinatti vs Concorrentes BR")
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=COLOR_HEADER)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(BRANDS))

# === Linha 2: Subtítulo data ===
ws.cell(row=2, column=1, value="Data: 2026-05-07 · ✓ = tem · ✗ = não tem · — = N/D · valores literais quando aplicável")
ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(BRANDS))

# === Linha 4: Cabeçalho de marcas ===
HEADER_ROW = 4
ws.cell(row=HEADER_ROW, column=1, value="Categoria").font = header_font
ws.cell(row=HEADER_ROW, column=2, value="Feature").font = header_font
ws.cell(row=HEADER_ROW, column=1).fill = PatternFill("solid", fgColor=COLOR_HEADER)
ws.cell(row=HEADER_ROW, column=2).fill = PatternFill("solid", fgColor=COLOR_HEADER)

for col_idx, brand in enumerate(BRANDS, start=3):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=brand)
    if brand == "Zerbinatti":
        cell.font = nosso_font
        cell.fill = PatternFill("solid", fgColor=COLOR_NOSSO)
    else:
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = center
    cell.border = border

ws.cell(row=HEADER_ROW, column=1).alignment = center
ws.cell(row=HEADER_ROW, column=2).alignment = center
ws.cell(row=HEADER_ROW, column=1).border = border
ws.cell(row=HEADER_ROW, column=2).border = border

# === Linha de URLs (logo abaixo do cabeçalho de marcas) ===
URL_ROW = HEADER_ROW + 1
ws.cell(row=URL_ROW, column=1, value="Link do site").font = Font(italic=True, size=9, color="666666")
ws.cell(row=URL_ROW, column=1).fill = PatternFill("solid", fgColor=COLOR_ALT)
ws.cell(row=URL_ROW, column=1).alignment = center
ws.cell(row=URL_ROW, column=1).border = border
ws.cell(row=URL_ROW, column=2, value="").fill = PatternFill("solid", fgColor=COLOR_ALT)
ws.cell(row=URL_ROW, column=2).border = border

for col_idx, brand in enumerate(BRANDS, start=3):
    url = BRAND_URLS[brand]
    # exibe domínio limpo, mantém URL completa no hyperlink
    display = url.replace("https://", "").replace("http://", "").rstrip("/")
    cell = ws.cell(row=URL_ROW, column=col_idx, value=display)
    cell.hyperlink = url
    cell.font = Font(size=8, color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    cell.fill = PatternFill("solid", fgColor=COLOR_ALT)

ws.row_dimensions[URL_ROW].height = 28

# === Linhas de dados ===
current_category = None
row_idx = URL_ROW + 1

for category, feature, values in ROWS:
    # Cabeçalho de categoria (uma linha grossa quando muda categoria)
    if category != current_category:
        cat_cell = ws.cell(row=row_idx, column=1, value=category)
        cat_cell.font = category_font
        cat_cell.fill = PatternFill("solid", fgColor=COLOR_CATEGORY)
        cat_cell.alignment = left
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2 + len(BRANDS))
        # aplicar fill em todas as cells mescladas
        for c in range(2, 3 + len(BRANDS)):
            ws.cell(row=row_idx, column=c).fill = PatternFill("solid", fgColor=COLOR_CATEGORY)
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        current_category = category

    # Linha de feature
    ws.cell(row=row_idx, column=1, value=category).font = Font(size=9, color="999999", italic=True)
    ws.cell(row=row_idx, column=1).alignment = left
    ws.cell(row=row_idx, column=2, value=feature).font = Font(size=10, bold=True)
    ws.cell(row=row_idx, column=2).alignment = left
    ws.cell(row=row_idx, column=1).border = border
    ws.cell(row=row_idx, column=2).border = border

    # Alternar cor de fundo da linha
    is_alt = (row_idx % 2 == 0)
    bg = COLOR_ALT if is_alt else "FFFFFF"
    ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=bg)
    ws.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor=bg)

    for col_idx, val in enumerate(values, start=3):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = center
        cell.border = border
        cell.font = Font(size=10)

        # Pintar fundo conforme conteúdo
        v = str(val).strip()
        if v == "✓":
            cell.fill = PatternFill("solid", fgColor=COLOR_OK)
            cell.font = Font(size=11, bold=True, color="1B5E20")
        elif v == "✗":
            cell.fill = PatternFill("solid", fgColor=COLOR_NO)
            cell.font = Font(size=11, bold=True, color="B71C1C")
        elif v == "—":
            cell.fill = PatternFill("solid", fgColor=COLOR_ND)
            cell.font = Font(size=10, color="666666")
        elif v.startswith("✓"):
            # ex: "✓ (84,75)"
            cell.fill = PatternFill("solid", fgColor=COLOR_OK)
            cell.font = Font(size=9, bold=False, color="1B5E20")
        elif v.startswith("✗"):
            cell.fill = PatternFill("solid", fgColor=COLOR_NO)
            cell.font = Font(size=9, color="B71C1C")
        else:
            cell.fill = PatternFill("solid", fgColor=bg)

        # Coluna do Zerbinatti recebe borda mais grossa
        if BRANDS[col_idx - 3] == "Zerbinatti":
            thick = Side(border_style="medium", color=COLOR_NOSSO)
            cell.border = Border(left=thick, right=thick, top=border.top, bottom=border.bottom)

    row_idx += 1

# === Larguras de coluna ===
ws.column_dimensions["A"].width = 22  # categoria
ws.column_dimensions["B"].width = 42  # feature
for col_idx, brand in enumerate(BRANDS, start=3):
    letter = get_column_letter(col_idx)
    ws.column_dimensions[letter].width = 18 if brand != "Zerbinatti" else 22

# Congelar painel após cabeçalho + linha de URL + colunas categoria/feature
ws.freeze_panes = "C6"

# === Aba 2: Resumo de scoring ===
ws2 = wb.create_sheet("Scoring por Marca")

# Calcular score: ✓ = 1 ponto, ✗ = 0, valor textual = 0.5, "—" = 0
def score_value(v):
    s = str(v).strip()
    if s == "✓" or s.startswith("✓"):
        return 1
    if s == "✗" or s.startswith("✗"):
        return 0
    if s == "—":
        return 0
    return 0.5  # texto livre = informação parcial

ws2.cell(row=1, column=1, value="Scoring por marca (✓=1, valor=0.5, ✗ ou —=0)").font = Font(bold=True, size=13)
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws2.cell(row=2, column=1, value="Calculado sobre todas as features booleanas/com valor da aba ‘Comparativo Master’.").font = Font(italic=True, size=9, color="666666")
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

ws2.cell(row=4, column=1, value="Marca").font = header_font
ws2.cell(row=4, column=2, value="Site").font = header_font
ws2.cell(row=4, column=3, value="Pontos").font = header_font
ws2.cell(row=4, column=4, value="Total possível").font = header_font
ws2.cell(row=4, column=5, value="% de cobertura").font = header_font
for c in range(1, 6):
    ws2.cell(row=4, column=c).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws2.cell(row=4, column=c).alignment = center
    ws2.cell(row=4, column=c).border = border

scores = []
total = len(ROWS)
for col_idx, brand in enumerate(BRANDS):
    s = sum(score_value(row[2][col_idx]) for row in ROWS)
    scores.append((brand, s, total, s / total * 100))

# Ordenar Zerbinatti primeiro, restante por score desc
scores_sorted = [scores[0]] + sorted(scores[1:], key=lambda x: -x[1])

for i, (brand, s, t, pct) in enumerate(scores_sorted, start=5):
    ws2.cell(row=i, column=1, value=brand)
    url = BRAND_URLS[brand]
    display = url.replace("https://", "").replace("http://", "").rstrip("/")
    link_cell = ws2.cell(row=i, column=2, value=display)
    link_cell.hyperlink = url
    link_cell.font = Font(color="0563C1", underline="single", size=10)
    ws2.cell(row=i, column=3, value=round(s, 1))
    ws2.cell(row=i, column=4, value=t)
    ws2.cell(row=i, column=5, value=f"{pct:.1f}%")
    for c in range(1, 6):
        ws2.cell(row=i, column=c).alignment = center
        ws2.cell(row=i, column=c).border = border
    if brand == "Zerbinatti":
        for c in range(1, 6):
            ws2.cell(row=i, column=c).fill = PatternFill("solid", fgColor=COLOR_NOSSO)
        # mantém link clicável mesmo com fundo escuro
        ws2.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=i, column=2).font = Font(color="FFE082", underline="single", size=10, bold=True)
        ws2.cell(row=i, column=3).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=i, column=4).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=i, column=5).font = Font(bold=True, color="FFFFFF")

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 38
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 18

# === Aba 3: Gaps Zerbinatti ===
ws3 = wb.create_sheet("Gaps do Zerbinatti")

ws3.cell(row=1, column=1, value="Features que o Zerbinatti NÃO tem (✗ ou —) mas concorrentes oferecem").font = Font(bold=True, size=13)
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

ws3.cell(row=3, column=1, value="Categoria").font = header_font
ws3.cell(row=3, column=2, value="Feature").font = header_font
ws3.cell(row=3, column=3, value="Quem tem (✓)").font = header_font
ws3.cell(row=3, column=4, value="Prioridade sugerida").font = header_font
for c in range(1, 5):
    ws3.cell(row=3, column=c).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws3.cell(row=3, column=c).alignment = center
    ws3.cell(row=3, column=c).border = border

# Prioridades manuais (alinhadas com seção 6.4 do relatório)
PRIORIDADES_ALTAS = {
    "Cupom primeira compra exibido",
    "Pix parcelado (Pagaleve etc)",
    "Multi-idioma (PT/EN/ES)",
    "Checkout em moeda estrangeira",
    "Envio internacional (DTC)",
    "Brinde de boas-vindas",
    "QR code de rastreabilidade no pacote",
    "Self-service B2B (checkout sem comercial)",
    "Vídeo embedded de preparo na ficha",
    "Filtro de reviews por descritor sensorial",
    "Programa de pontos",
    "Apple Pay / Google Pay",
    "Premiações Cup of Excellence comunicadas",
    "Selo B Corp / sustentabilidade ESG",
    "Embalagem proprietária (ex: Penta Box)",
}
PRIORIDADES_MEDIAS = {
    "Kits / Presentes",
    "Acessórios (moedor, bule, xícara)",
    "Cápsulas",
    "1kg",
    "Sample / 50g",
    "Moagem por método (V60, Moka, Prensa, Espresso)",
    "Programa fidelidade / cashback",
    "Cursos pagos / escola",
    "Boleto",
}

gap_row = 4
for category, feature, values in ROWS:
    nosso = str(values[0]).strip()
    if nosso == "✓" or nosso.startswith("✓"):
        continue
    # quem tem
    quem = []
    for i, val in enumerate(values[1:], start=1):
        v = str(val).strip()
        if v == "✓" or v.startswith("✓"):
            quem.append(BRANDS[i])
    if not quem:
        continue  # se ninguém tem, não é gap competitivo
    if feature in PRIORIDADES_ALTAS:
        prio = "🔴 ALTA"
    elif feature in PRIORIDADES_MEDIAS:
        prio = "🟡 MÉDIA"
    else:
        prio = "🟢 BAIXA"

    ws3.cell(row=gap_row, column=1, value=category)
    ws3.cell(row=gap_row, column=2, value=feature).font = Font(bold=True)
    ws3.cell(row=gap_row, column=3, value=", ".join(quem))
    ws3.cell(row=gap_row, column=4, value=prio)
    for c in range(1, 5):
        ws3.cell(row=gap_row, column=c).alignment = left
        ws3.cell(row=gap_row, column=c).border = border
    if "ALTA" in prio:
        ws3.cell(row=gap_row, column=4).fill = PatternFill("solid", fgColor="FFCDD2")
    elif "MÉDIA" in prio:
        ws3.cell(row=gap_row, column=4).fill = PatternFill("solid", fgColor="FFF9C4")
    else:
        ws3.cell(row=gap_row, column=4).fill = PatternFill("solid", fgColor="C8E6C9")
    gap_row += 1

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 42
ws3.column_dimensions["C"].width = 50
ws3.column_dimensions["D"].width = 18
ws3.freeze_panes = "A4"

# === Aba 4: Vantagens Zerbinatti ===
ws4 = wb.create_sheet("Vantagens Zerbinatti")
ws4.cell(row=1, column=1, value="Features que o Zerbinatti TEM e a maioria dos concorrentes NÃO tem").font = Font(bold=True, size=13)
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

ws4.cell(row=3, column=1, value="Categoria").font = header_font
ws4.cell(row=3, column=2, value="Feature").font = header_font
ws4.cell(row=3, column=3, value="Concorrentes que têm").font = header_font
ws4.cell(row=3, column=4, value="Diferenciação").font = header_font
for c in range(1, 5):
    ws4.cell(row=3, column=c).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws4.cell(row=3, column=c).alignment = center

adv_row = 4
for category, feature, values in ROWS:
    nosso = str(values[0]).strip()
    if not (nosso == "✓" or nosso.startswith("✓")):
        continue
    quem = []
    for i, val in enumerate(values[1:], start=1):
        v = str(val).strip()
        if v == "✓" or v.startswith("✓"):
            quem.append(BRANDS[i])
    n = len(quem)
    if n >= 6:
        continue  # commodity, todo mundo tem
    if n == 0:
        diff = "🏆 ÚNICO no mercado"
    elif n <= 2:
        diff = "💎 Diferenciado (poucos têm)"
    else:
        diff = "✨ Acima da média"
    ws4.cell(row=adv_row, column=1, value=category)
    ws4.cell(row=adv_row, column=2, value=feature).font = Font(bold=True)
    ws4.cell(row=adv_row, column=3, value=", ".join(quem) if quem else "(nenhum)")
    ws4.cell(row=adv_row, column=4, value=diff)
    if "ÚNICO" in diff:
        ws4.cell(row=adv_row, column=4).fill = PatternFill("solid", fgColor="FFD700")
    elif "Diferenciado" in diff:
        ws4.cell(row=adv_row, column=4).fill = PatternFill("solid", fgColor="C8E6C9")
    else:
        ws4.cell(row=adv_row, column=4).fill = PatternFill("solid", fgColor="E1F5FE")
    for c in range(1, 5):
        ws4.cell(row=adv_row, column=c).alignment = left
    adv_row += 1

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 42
ws4.column_dimensions["C"].width = 50
ws4.column_dimensions["D"].width = 26
ws4.freeze_panes = "A4"

# === Salvar ===
wb.save(OUT)
print(f"Salvo em: {OUT}")
print(f"Total features: {len(ROWS)}")
print(f"Marcas comparadas: {len(BRANDS)}")
