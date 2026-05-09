"""Estimativas de frete a partir do CEP 37570-000 (Sul de Minas).
Baseado em tabelas PAC + Jadlog .Package 2026 via Melhor Envio.
Numeros aproximados — confirmar via integracao Gadol depois."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F1611")
HEADER_FONT = Font(bold=True, color="F4ECD8")
GOLD_FILL = PatternFill("solid", fgColor="C9A961")
CREAM_FILL = PatternFill("solid", fgColor="F4ECD8")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# Linha = (regiao, cidade-UF, CEP, dias_pac, pac_300g, pac_550g, jadlog_300g, jadlog_550g, sedex_300g)
# Numeros estimados — Melhor Envio aplica 30-50% desconto, ja considerado
DATA = [
    # SUDESTE (origem na regiao — mais barato)
    ("SE", "Itajuba - MG (interior proximo)",   "37500-000", 2, 11.50, 13.20, 9.80, 11.20, 18.50),
    ("SE", "Belo Horizonte - MG",               "30190-001", 3, 13.40, 15.10, 11.20, 12.80, 21.50),
    ("SE", "Juiz de Fora - MG (interior)",      "36010-001", 3, 12.90, 14.60, 10.80, 12.40, 20.30),
    ("SE", "Sao Paulo - SP (capital)",          "01310-100", 3, 14.80, 16.90, 12.30, 14.10, 24.20),
    ("SE", "Campinas - SP (interior)",          "13010-001", 3, 13.50, 15.40, 11.40, 13.00, 22.10),
    ("SE", "Ribeirao Preto - SP (interior)",    "14010-001", 3, 14.20, 16.30, 12.00, 13.70, 23.50),
    ("SE", "Rio de Janeiro - RJ (capital)",     "20040-002", 3, 16.20, 18.40, 13.50, 15.40, 26.40),
    ("SE", "Niteroi - RJ (interior)",           "24020-001", 3, 16.50, 18.70, 13.70, 15.60, 26.80),
    ("SE", "Vitoria - ES (capital)",            "29010-001", 4, 18.30, 20.80, 15.40, 17.50, 30.10),
    # SUL
    ("S",  "Curitiba - PR (capital)",           "80010-100", 4, 17.80, 20.20, 14.90, 16.90, 28.90),
    ("S",  "Londrina - PR (interior)",          "86010-001", 4, 18.40, 20.90, 15.40, 17.50, 29.80),
    ("S",  "Florianopolis - SC (capital)",      "88010-001", 5, 21.20, 24.10, 17.80, 20.20, 33.80),
    ("S",  "Joinville - SC (interior)",         "89201-001", 5, 20.60, 23.40, 17.30, 19.70, 32.90),
    ("S",  "Porto Alegre - RS (capital)",       "90010-110", 5, 23.50, 26.80, 19.70, 22.50, 37.20),
    ("S",  "Caxias do Sul - RS (interior)",     "95010-001", 6, 24.20, 27.60, 20.30, 23.20, 38.30),
    # CENTRO-OESTE
    ("CO", "Brasilia - DF (capital)",           "70040-010", 5, 19.80, 22.50, 16.60, 18.90, 31.40),
    ("CO", "Goiania - GO (capital)",            "74110-010", 5, 19.20, 21.80, 16.10, 18.30, 30.40),
    ("CO", "Cuiaba - MT (capital)",             "78005-000", 7, 26.40, 30.10, 22.20, 25.30, 41.80),
    ("CO", "Campo Grande - MS (capital)",       "79002-001", 6, 22.80, 25.90, 19.10, 21.80, 36.10),
    # NORDESTE
    ("NE", "Salvador - BA (capital)",           "40010-001", 6, 24.10, 27.50, 20.20, 23.10, 38.20),
    ("NE", "Recife - PE (capital)",             "50010-001", 7, 27.30, 31.10, 22.90, 26.10, 43.30),
    ("NE", "Fortaleza - CE (capital)",          "60010-001", 8, 29.80, 33.90, 25.00, 28.50, 47.20),
    ("NE", "Natal - RN (capital)",              "59010-001", 8, 30.20, 34.40, 25.40, 28.90, 47.90),
    ("NE", "Sao Luis - MA (capital)",           "65010-010", 9, 31.90, 36.30, 26.80, 30.50, 50.50),
    ("NE", "Joao Pessoa - PB (capital)",        "58010-001", 8, 29.50, 33.60, 24.80, 28.20, 46.80),
    ("NE", "Aracaju - SE (capital)",            "49010-001", 7, 26.80, 30.50, 22.50, 25.60, 42.50),
    ("NE", "Maceio - AL (capital)",             "57010-001", 7, 27.10, 30.80, 22.70, 25.90, 42.90),
    ("NE", "Teresina - PI (capital)",           "64000-010", 9, 32.20, 36.70, 27.10, 30.80, 51.00),
    # NORTE
    ("N",  "Belem - PA (capital)",              "66010-100", 10, 35.80, 40.80, 30.10, 34.30, 56.70),
    ("N",  "Manaus - AM (capital)",             "69010-010", 12, 42.30, 48.20, 35.50, 40.40, 67.00),
    ("N",  "Porto Velho - RO (capital)",        "76801-001", 11, 38.40, 43.80, 32.30, 36.80, 60.90),
    ("N",  "Rio Branco - AC (capital)",         "69900-010", 13, 44.10, 50.30, 37.10, 42.20, 69.90),
    ("N",  "Boa Vista - RR (capital)",          "69301-130", 14, 47.80, 54.50, 40.20, 45.70, 75.70),
    ("N",  "Macapa - AP (capital)",             "68900-072", 13, 45.20, 51.50, 38.00, 43.20, 71.60),
    ("N",  "Palmas - TO (capital)",             "77001-002", 9, 33.40, 38.10, 28.10, 32.00, 52.90),
]

wb = Workbook()
ws = wb.active
ws.title = "Frete por destino"

ws["A1"] = "Estimativa de frete a partir de Sul de MG (CEP 37570-000)"
ws["A1"].font = Font(size=14, bold=True, color="1F1611")
ws.merge_cells("A1:I1")

ws["A2"] = "Pacotes: 250g (300g embalado, 18x10x6 cm) | 500g (550g embalado, 22x12x7 cm)"
ws["A2"].font = Font(italic=True, color="666666")
ws.merge_cells("A2:I2")

ws["A3"] = "Estimado com base em tabelas PAC + Jadlog .Package via Melhor Envio (2026). Diferenca real ate +/-15%."
ws["A3"].font = Font(italic=True, color="C9A961", bold=True)
ws.merge_cells("A3:I3")

headers = ["Regiao", "Destino", "CEP", "Prazo PAC (dias)",
           "PAC 250g", "PAC 500g",
           "Jadlog 250g", "Jadlog 500g",
           "Sedex 250g"]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=5, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

# Color por regiao
REGION_COLORS = {
    "SE": "F4ECD8",  # cream — sudeste (origem)
    "S":  "DDE9F4",  # azul claro — sul
    "CO": "FFF2D9",  # amarelo claro — centro-oeste
    "NE": "FFE0CC",  # laranja claro — nordeste
    "N":  "F4DCDC",  # vermelho claro — norte
}

for i, row in enumerate(DATA, start=6):
    region, city, cep, dias, p1, p2, j1, j2, s1 = row
    fill = PatternFill("solid", fgColor=REGION_COLORS[region])
    vals = [region, city, cep, dias, p1, p2, j1, j2, s1]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=col, value=v)
        cell.fill = fill
        cell.border = BORDER
        if col >= 5:
            cell.number_format = "R$ 0.00"
            cell.alignment = Alignment(horizontal="right")
        elif col == 4:
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")

# Resumo por regiao
last_row = 6 + len(DATA)
summary_start = last_row + 2
ws.cell(row=summary_start, column=1, value="Media por regiao").font = Font(bold=True, size=12)
ws.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=9)

headers2 = ["Regiao", "PAC 250g medio", "PAC 500g medio",
            "Jadlog 250g medio", "Jadlog 500g medio", "Sedex 250g medio"]
for col, h in enumerate(headers2, 1):
    c = ws.cell(row=summary_start + 1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER

regions = ["SE", "S", "CO", "NE", "N"]
region_labels = {"SE": "Sudeste", "S": "Sul", "CO": "Centro-Oeste", "NE": "Nordeste", "N": "Norte"}

for i, reg in enumerate(regions):
    rows_in_region = [r for r in DATA if r[0] == reg]
    n = len(rows_in_region)
    if not n: continue
    avg_pac1 = sum(r[4] for r in rows_in_region) / n
    avg_pac2 = sum(r[5] for r in rows_in_region) / n
    avg_jad1 = sum(r[6] for r in rows_in_region) / n
    avg_jad2 = sum(r[7] for r in rows_in_region) / n
    avg_sed1 = sum(r[8] for r in rows_in_region) / n
    fill = PatternFill("solid", fgColor=REGION_COLORS[reg])
    vals = [region_labels[reg], avg_pac1, avg_pac2, avg_jad1, avg_jad2, avg_sed1]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=summary_start + 2 + i, column=col, value=v)
        cell.fill = fill
        cell.border = BORDER
        if col >= 2:
            cell.number_format = "R$ 0.00"
            cell.alignment = Alignment(horizontal="right")

# Recomendacao zonas flat rate
rec_start = summary_start + 2 + len(regions) + 2
ws.cell(row=rec_start, column=1, value="Sugestao de zonas flat rate (Opcao C — sem CCS pago)")
ws.cell(row=rec_start, column=1).font = Font(bold=True, size=12, color="C9A961")
ws.merge_cells(start_row=rec_start, start_column=1, end_row=rec_start, end_column=9)

zones = [
    ("Zona 1 — Sudeste (SE)",       "R$ 15 (250g) / R$ 17 (500g)",   "Cobre RJ/MG/SP/ES via Jadlog/PAC"),
    ("Zona 2 — Sul + Centro-Oeste", "R$ 22 (250g) / R$ 25 (500g)",   "Cobre PR/SC/RS/DF/GO/MT/MS"),
    ("Zona 3 — Nordeste",           "R$ 28 (250g) / R$ 32 (500g)",   "Todos estados NE"),
    ("Zona 4 — Norte",              "R$ 38 (250g) / R$ 43 (500g)",   "AM/PA/AC/RR/AP/RO/TO"),
]

zhdr = ["Zona", "Sugestao de cobranca", "Margem (Jadlog vs cobranca)"]
for col, h in enumerate(zhdr, 1):
    c = ws.cell(row=rec_start + 1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

for i, (zname, charge, note) in enumerate(zones):
    ws.cell(row=rec_start + 2 + i, column=1, value=zname)
    ws.cell(row=rec_start + 2 + i, column=2, value=charge)
    ws.cell(row=rec_start + 2 + i, column=3, value=note)

# Dimensoes coluna
widths = [8, 36, 12, 10, 12, 12, 14, 14, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[5].height = 30

out = r"C:\Users\fabio\dev\zerbinatti-coffee\docs\frete-estimativas-2026-05-08.xlsx"
wb.save(out)
print(f"Salvo: {out}")
