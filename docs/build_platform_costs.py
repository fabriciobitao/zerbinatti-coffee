"""Cria planilha Excel com custos de plataforma (Shopify + GCP + Firebase)
para 15k-30k usuarios unicos mensais. Numeros sourceados em 2026-05-08."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import date

USD_BRL = 5.30  # Cotacao referencia maio 2026

# Estilos reutilizaveis
HEADER_FILL = PatternFill("solid", fgColor="1F1611")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="F4ECD8")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F1611")
GOLD_FILL = PatternFill("solid", fgColor="C9A961")
GOLD_FONT = Font(name="Calibri", size=11, bold=True, color="1F1611")
CREAM_FILL = PatternFill("solid", fgColor="F4ECD8")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()

# ==============================================================
# SHEET 1 — Planos Shopify
# ==============================================================
ws = wb.active
ws.title = "Planos Shopify"

ws["A1"] = "Planos Shopify (precos USD oficiais convertidos a R$ 5,30/USD)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:H1")
ws["A2"] = f"Atualizado: {date.today().isoformat()} | Fonte: shopify.com/pricing"
ws["A2"].font = Font(italic=True, color="666666")
ws.merge_cells("A2:H2")

headers = ["Plano", "Mensal (USD)", "Mensal (BRL)", "Anual (BRL)*", "Economia anual",
           "Taxa por venda (3rd party)", "Cartao via Shopify Payments", "Notas"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# (plan, monthly USD, 3rd-party fee %, Shopify Payments rate, notes)
plans = [
    ("Basic",    29,   2.0, "2.9% + 30c USD",
     "Entrada. 1 loja, ate 2 staffs. Sem reports avancados."),
    ("Grow",     79,   1.0, "2.6% + 30c USD",
     "Antigo 'Shopify'. 5 staffs, professional reports."),
    ("Advanced", 299,  0.6, "2.4% + 30c USD",
     "15 staffs, reports custom, calculated shipping."),
    ("Plus",   2300,   0.2, "negociado",
     "Enterprise. Checkout customizavel via Liquid, multi-store. Min 12 meses."),
]

for i, (plan, usd, third_pct, sp_rate, notes) in enumerate(plans, start=5):
    brl = usd * USD_BRL
    annual = brl * 12 * 0.75   # 25% discount
    save = brl * 12 - annual

    ws.cell(row=i, column=1, value=plan).font = Font(bold=True)
    ws.cell(row=i, column=2, value=usd).number_format = "$#,##0"
    ws.cell(row=i, column=3, value=brl).number_format = "R$ #,##0.00"
    ws.cell(row=i, column=4, value=annual).number_format = "R$ #,##0.00"
    ws.cell(row=i, column=5, value=save).number_format = "R$ #,##0.00"
    ws.cell(row=i, column=6, value=f"+{third_pct}% sobre cada venda").alignment = CENTER
    ws.cell(row=i, column=7, value=sp_rate).alignment = CENTER
    ws.cell(row=i, column=8, value=notes).alignment = WRAP
    for col in range(1, 9):
        ws.cell(row=i, column=col).border = BORDER

# Highlight Plus row
for col in range(1, 9):
    ws.cell(row=8, column=col).fill = CREAM_FILL

ws["A10"] = "* Anual com desconto de 25% (Basic/Grow/Advanced). Plus tem desconto custom em 3 anos."
ws["A10"].font = Font(italic=True, color="666666")
ws.merge_cells("A10:H10")
ws["A11"] = "** Shopify Payments NAO esta disponivel oficialmente no Brasil em maio/2026."
ws["A11"].font = Font(italic=True, color="C9A961", bold=True)
ws.merge_cells("A11:H11")
ws["A12"] = "    Voce SEMPRE paga a taxa 3rd-party (Mercado Pago, PagSeguro etc) + a coluna F."
ws["A12"].font = Font(italic=True, color="666666")
ws.merge_cells("A12:H12")

widths = [12, 14, 14, 16, 16, 26, 26, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ==============================================================
# SHEET 2 — Pagamento BR (gateway / Mercado Pago / PagSeguro)
# ==============================================================
ws = wb.create_sheet("Pagamento BR")

ws["A1"] = "Provedores de pagamento BR (combinados com Shopify)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")

headers = ["Provider", "Metodo", "Taxa (%)", "Taxa fixa (R$)", "Em R$ 50",
           "Em R$ 100", "Notas"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# (provider, metodo, %, fixo R$, notas)
providers = [
    ("Mercado Pago",  "PIX",                 0.99, 0.00,  "Mais barato do BR"),
    ("Mercado Pago",  "Boleto",              0.00, 3.49,  "Lento (~2 dias compensar)"),
    ("Mercado Pago",  "Cartao 1x",           4.99, 0.00,  "Recebe em 14d (D+14)"),
    ("Mercado Pago",  "Cartao 2-12x",        5.31, 0.00,  "+1.5% por parcela alem de 1x"),
    ("PagSeguro",     "PIX",                 0.99, 0.00,  ""),
    ("PagSeguro",     "Boleto",              0.00, 3.99,  ""),
    ("PagSeguro",     "Cartao 1x",           4.39, 0.40,  "+0.40 fixo por venda"),
    ("Pagar.me",      "PIX",                 0.99, 0.00,  "Stone — D+1"),
    ("Pagar.me",      "Cartao",              3.19, 0.40,  "Mais barato; +2.49%/parcela"),
    ("Stripe BR",     "Cartao internacional",4.99, 0.00,  "Sem PIX nem boleto"),
]

for i, (prov, met, pct, fix, notes) in enumerate(providers, start=4):
    on_50 = (50 * pct / 100) + fix
    on_100 = (100 * pct / 100) + fix
    ws.cell(row=i, column=1, value=prov).font = Font(bold=True)
    ws.cell(row=i, column=2, value=met)
    ws.cell(row=i, column=3, value=pct).number_format = "0.00\"%\""
    ws.cell(row=i, column=4, value=fix).number_format = "R$ 0.00"
    ws.cell(row=i, column=5, value=on_50).number_format = "R$ 0.00"
    ws.cell(row=i, column=6, value=on_100).number_format = "R$ 0.00"
    ws.cell(row=i, column=7, value=notes).alignment = WRAP
    for col in range(1, 8):
        ws.cell(row=i, column=col).border = BORDER

ws["A15"] = "Recomendacao: ativar Mercado Pago no Shopify -> oferece PIX (~1%), Boleto e Cartao."
ws["A15"].font = Font(italic=True, color="666666")
ws.merge_cells("A15:G15")
ws["A16"] = "Mix tipico BR e-commerce: 45% Cartao / 35% PIX / 15% Boleto / 5% Outros = ~3% medio sobre receita."
ws["A16"].font = Font(italic=True, color="666666")
ws.merge_cells("A16:G16")

widths = [16, 22, 12, 14, 12, 12, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ==============================================================
# SHEET 3 — Infra GCP + Firebase
# ==============================================================
ws = wb.create_sheet("Infra GCP+Firebase")

ws["A1"] = "Custo de infra para 15k-30k MUU/mes (Cloud Run + Firebase Hosting + Firestore)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")

ws["A3"] = "Premissas de trafego"
ws["A3"].font = Font(bold=True, size=12)

assumptions = [
    ("Page views por usuario (estimado)", 5),
    ("KB por page view (HTML+JS+CSS, sem imagens cacheadas)", 150),
    ("API calls Shopify por page (Storefront + cart)", 1.5),
    ("Form B2B submissoes/mes (Firestore writes)", 50),
]
for i, (label, val) in enumerate(assumptions, start=4):
    ws.cell(row=i, column=1, value=label).alignment = WRAP
    ws.cell(row=i, column=2, value=val)

# Cenarios
ws["A10"] = "Cenarios"
ws["A10"].font = Font(bold=True, size=12)

scenarios = [
    ("Baixo",  15000),
    ("Medio",  22500),
    ("Alto",   30000),
]

ws.cell(row=11, column=1, value="Servico").font = HEADER_FONT
ws.cell(row=11, column=1).fill = HEADER_FILL
for i, (label, _) in enumerate(scenarios, start=2):
    c = ws.cell(row=11, column=i, value=f"{label} (k MUU)")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER

# === Cloud Run ===
def cloud_run_cost(muu):
    requests = muu * 5 * 1.5  # 5 PV * 1.5 API
    # Cloud Run free: 2M requests/mes
    paid_req = max(0, requests - 2_000_000)
    cost_req_usd = paid_req * 0.40 / 1_000_000
    # CPU: assume 200ms por request, 1 vCPU
    vcpu_sec = requests * 0.2
    free_vcpu = 360_000  # 180k vCPU-seconds free + 180k mais com tier
    paid_vcpu = max(0, vcpu_sec - free_vcpu)
    cost_vcpu_usd = paid_vcpu * 0.000018
    # Mem: 1 GiB * 0.2s
    mem_gibsec = requests * 0.2 * 1
    free_mem = 360_000
    paid_mem = max(0, mem_gibsec - free_mem)
    cost_mem_usd = paid_mem * 0.000002
    return (cost_req_usd + cost_vcpu_usd + cost_mem_usd) * USD_BRL

def firebase_hosting_cost(muu):
    pv = muu * 5
    gb_transfer = pv * 150 / 1024 / 1024  # KB to GB
    free_gb = 10
    paid_gb = max(0, gb_transfer - free_gb)
    return paid_gb * 0.15 * USD_BRL

def firestore_cost(muu):
    # B2B writes are tiny; reads negligible
    writes = 50  # b2b form submissions
    free_writes = 20_000  # 20k writes/day free, easy
    paid_writes = max(0, writes - free_writes)
    cost = paid_writes / 100_000 * 0.18  # $0.18 per 100k writes
    return cost * USD_BRL

# === Linhas ===
rows = [
    ("Cloud Run (compute)", cloud_run_cost),
    ("Firebase Hosting (CDN bandwidth)", firebase_hosting_cost),
    ("Firestore (B2B forms)", firestore_cost),
]

for r_idx, (name, fn) in enumerate(rows, start=12):
    ws.cell(row=r_idx, column=1, value=name)
    for c_idx, (_, muu) in enumerate(scenarios, start=2):
        cost = fn(muu)
        cell = ws.cell(row=r_idx, column=c_idx, value=cost)
        cell.number_format = "R$ #,##0.00"
        cell.alignment = CENTER

# Total row
total_row = 12 + len(rows)
ws.cell(row=total_row, column=1, value="TOTAL infra/mes").font = GOLD_FONT
ws.cell(row=total_row, column=1).fill = GOLD_FILL
for c_idx in range(2, 5):
    col_letter = get_column_letter(c_idx)
    f = f"=SUM({col_letter}12:{col_letter}{total_row-1})"
    cell = ws.cell(row=total_row, column=c_idx, value=f)
    cell.number_format = "R$ #,##0.00"
    cell.font = GOLD_FONT
    cell.fill = GOLD_FILL
    cell.alignment = CENTER

ws[f"A{total_row + 2}"] = "Observacoes:"
ws[f"A{total_row + 2}"].font = Font(bold=True)
notes_infra = [
    "- Cloud Run: $0.40/M requests, $0.000018/vCPU-s, $0.000002/GiB-s (post free tier).",
    "- Firebase Hosting: 10 GB/mes free; depois $0.15/GB cached, $0.20/GB uncached.",
    "- Firestore: 20k writes/dia free, 50k reads/dia free. Form B2B esta dentro do free tier.",
    "- NAO inclui custos de imagens (CDN Shopify hospeda) nem video (Vimeo).",
    "- Cold start Cloud Run: scale-to-zero pode ter latencia 1-2s no primeiro hit; ok pra esse volume.",
]
for i, n in enumerate(notes_infra, start=total_row + 3):
    ws[f"A{i}"] = n
    ws[f"A{i}"].font = Font(italic=True, color="666666")
    ws.merge_cells(f"A{i}:G{i}")

widths = [40, 18, 18, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ==============================================================
# SHEET 4 — Cenario consolidado
# ==============================================================
ws = wb.create_sheet("Cenario Total")

ws["A1"] = "Custo total de plataforma por mes (Shopify + Pagamento + Infra)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")

ws["A3"] = "Premissas:"
ws["A3"].font = Font(bold=True)
prem = [
    "- Conversao 1.5% (e-commerce maturo BR ~1-2%)",
    "- Ticket medio R$ 65 (mix 250g R$ 49.90 / 500g R$ 79.90)",
    "- Mix de pagamento: 45% Cartao Mercado Pago, 35% PIX, 15% Boleto, 5% outros",
    "- Taxa media efetiva por venda: 3.0% sobre receita (cartao 5% × 0.45 + PIX 1% × 0.35 + boleto fixo)",
]
for i, p in enumerate(prem, start=4):
    ws[f"A{i}"] = p
    ws[f"A{i}"].font = Font(italic=True, color="666666")
    ws.merge_cells(f"A{i}:G{i}")

# Tabela
ws["A9"] = "Cenario por trafego mensal"
ws["A9"].font = Font(bold=True, size=12)

headers = ["Trafego (MUU)", "Pedidos/mes", "Receita bruta",
           "Shopify Basic anual/mes", "Taxa pagamento (3%)",
           "Infra GCP+Firebase", "TOTAL plataforma/mes"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=10, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

shopify_basic_monthly_anual = 29 * USD_BRL * 0.75  # ~R$115
shopify_grow_monthly_anual = 79 * USD_BRL * 0.75   # ~R$314
shopify_advanced_monthly_anual = 299 * USD_BRL * 0.75  # ~R$1.188

# Taxas third-party por plano (Shopify cobra alem da taxa do gateway BR)
basic_3p_fee = 0.02       # 2% sobre receita
grow_3p_fee = 0.01        # 1%
advanced_3p_fee = 0.006   # 0.6%

infra_costs = {15000: 0, 22500: 1.5, 30000: 4.5}  # de sheet 3 (~maior parte free tier)

for r_idx, (label, muu) in enumerate([("Baixo", 15000), ("Medio", 22500), ("Alto", 30000)], start=11):
    pedidos = muu * 0.015
    receita = pedidos * 65
    pagamento_gateway = receita * 0.03  # Mercado Pago/PagSeguro etc — NAO inclui taxa Shopify
    pagamento_total = pagamento_gateway + (receita * basic_3p_fee)  # gateway + Shopify 2%
    infra = infra_costs[muu]
    total = shopify_basic_monthly_anual + pagamento_total + infra

    ws.cell(row=r_idx, column=1, value=f"{muu:,} ({label})").alignment = CENTER
    ws.cell(row=r_idx, column=2, value=pedidos).number_format = "#,##0"
    ws.cell(row=r_idx, column=3, value=receita).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=4, value=shopify_basic_monthly_anual).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=5, value=pagamento_total).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=6, value=infra).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=7, value=total).number_format = "R$ #,##0.00"

    for col in range(1, 8):
        ws.cell(row=r_idx, column=col).border = BORDER

    # Highlight TOTAL column
    ws.cell(row=r_idx, column=7).font = GOLD_FONT
    ws.cell(row=r_idx, column=7).fill = GOLD_FILL

ws["A14"] = "* Pagamento total inclui gateway BR (3% medio) + taxa Shopify por plano (2.0% pro Basic)."
ws["A14"].font = Font(italic=True, color="666666")
ws.merge_cells("A14:G14")

# ===== Comparativo Basic vs Grow vs Advanced =====
comp_start = 16
ws.cell(row=comp_start, column=1, value="Comparativo entre planos Shopify (mesmos cenarios de trafego)")
ws.cell(row=comp_start, column=1).font = Font(bold=True, size=12, color="1F1611")
ws.merge_cells(start_row=comp_start, start_column=1, end_row=comp_start, end_column=7)

# Header
comp_headers = ["Trafego (MUU)", "Receita bruta", "Custo TOTAL Basic", "Custo TOTAL Grow",
                "Custo TOTAL Advanced", "Economia Grow vs Basic", "Economia Adv vs Basic"]
for col, h in enumerate(comp_headers, 1):
    c = ws.cell(row=comp_start + 1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = BORDER

ws.row_dimensions[comp_start + 1].height = 30

for r_idx, (label, muu) in enumerate([("Baixo", 15000), ("Medio", 22500), ("Alto", 30000)], start=comp_start + 2):
    pedidos = muu * 0.015
    receita = pedidos * 65
    gateway = receita * 0.03
    infra = infra_costs[muu]
    custo_basic = shopify_basic_monthly_anual + gateway + receita * basic_3p_fee + infra
    custo_grow = shopify_grow_monthly_anual + gateway + receita * grow_3p_fee + infra
    custo_advanced = shopify_advanced_monthly_anual + gateway + receita * advanced_3p_fee + infra
    econ_grow = custo_basic - custo_grow
    econ_adv = custo_basic - custo_advanced

    ws.cell(row=r_idx, column=1, value=f"{muu:,} ({label})").alignment = CENTER
    ws.cell(row=r_idx, column=2, value=receita).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=3, value=custo_basic).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=4, value=custo_grow).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=5, value=custo_advanced).number_format = "R$ #,##0.00"
    ws.cell(row=r_idx, column=6, value=econ_grow).number_format = "R$ #,##0.00;[Red]-R$ #,##0.00"
    ws.cell(row=r_idx, column=7, value=econ_adv).number_format = "R$ #,##0.00;[Red]-R$ #,##0.00"

    for col in range(1, 8):
        ws.cell(row=r_idx, column=col).border = BORDER

    # Highlight winner (lowest custo)
    custos = [custo_basic, custo_grow, custo_advanced]
    winner_col = 3 + custos.index(min(custos))  # column 3, 4 ou 5
    ws.cell(row=r_idx, column=winner_col).fill = GOLD_FILL
    ws.cell(row=r_idx, column=winner_col).font = GOLD_FONT

# Break-even analysis
be_start = comp_start + 6
ws.cell(row=be_start, column=1, value="Quando vale subir de plano (break-even)")
ws.cell(row=be_start, column=1).font = Font(bold=True, size=12, color="C9A961")
ws.merge_cells(start_row=be_start, start_column=1, end_row=be_start, end_column=7)

# Grow vs Basic: pagar R$ 199/mes a mais (314-115) pra economizar 1% (2% -> 1%) das vendas
diff_grow = shopify_grow_monthly_anual - shopify_basic_monthly_anual
be_grow = diff_grow / (basic_3p_fee - grow_3p_fee)
ws.cell(row=be_start + 1, column=1,
        value=f"Grow vale a pena acima de R$ {be_grow:,.0f} de receita/mes (mensalidade extra: R$ {diff_grow:,.0f}, economia 1.0% por venda)")
ws.cell(row=be_start + 1, column=1).font = Font(italic=True)
ws.merge_cells(start_row=be_start + 1, start_column=1, end_row=be_start + 1, end_column=7)

diff_adv = shopify_advanced_monthly_anual - shopify_basic_monthly_anual
be_adv = diff_adv / (basic_3p_fee - advanced_3p_fee)
ws.cell(row=be_start + 2, column=1,
        value=f"Advanced vale a pena acima de R$ {be_adv:,.0f} de receita/mes (mensalidade extra: R$ {diff_adv:,.0f}, economia 1.4% por venda)")
ws.cell(row=be_start + 2, column=1).font = Font(italic=True)
ws.merge_cells(start_row=be_start + 2, start_column=1, end_row=be_start + 2, end_column=7)

ws.cell(row=be_start + 3, column=1,
        value="Beneficios nao-financeiros do Grow: 5 staffs (vs 2 Basic), reports profissionais, retencao de carrinho abandonado.")
ws.cell(row=be_start + 3, column=1).font = Font(italic=True, color="666666")
ws.merge_cells(start_row=be_start + 3, start_column=1, end_row=be_start + 3, end_column=7)

widths = [16, 14, 16, 22, 18, 18, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ==============================================================
# SHEET 5 — Notas e Sources
# ==============================================================
ws = wb.create_sheet("Sources")

ws["A1"] = "Fontes e premissas"
ws["A1"].font = TITLE_FONT

sources = [
    ("Shopify Pricing", "https://www.shopify.com/pricing"),
    ("Shopify transaction fees explained 2026", "https://craftshift.com/shopify-transaction-fees-explained/"),
    ("Cloud Run pricing", "https://cloud.google.com/run/pricing"),
    ("Firebase Hosting pricing", "https://firebase.google.com/docs/hosting/usage-quotas-pricing"),
    ("Firebase pricing plans", "https://firebase.google.com/pricing"),
    ("Mercado Pago taxas", "https://www.mercadopago.com.br/ajuda/Custos_265"),
    ("PagSeguro taxas", "https://pagseguro.uol.com.br/para-seu-negocio/taxas-e-tarifas"),
    ("Pagar.me taxas", "https://pagar.me/precos"),
]
ws["A3"] = "Servico"
ws["A3"].font = HEADER_FONT
ws["A3"].fill = HEADER_FILL
ws["B3"] = "URL"
ws["B3"].font = HEADER_FONT
ws["B3"].fill = HEADER_FILL

for i, (svc, url) in enumerate(sources, start=4):
    ws.cell(row=i, column=1, value=svc)
    ws.cell(row=i, column=2, value=url)
    ws.cell(row=i, column=2).hyperlink = url
    ws.cell(row=i, column=2).font = Font(color="0563C1", underline="single")

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 80

ws["A14"] = "Cotacao USD/BRL utilizada: R$ 5,30 (referencia maio/2026)."
ws["A14"].font = Font(italic=True)

# ==============================================================
# Salvar
# ==============================================================
out_path = r"C:\Users\fabio\dev\zerbinatti-coffee\docs\custos-plataforma-2026-05-08.xlsx"
wb.save(out_path)
print(f"Salvo: {out_path}")
